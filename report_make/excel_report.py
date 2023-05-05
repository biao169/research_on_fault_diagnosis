import string
import copy
import pandas as pd
import openpyxl as oxl
from openpyxl.styles import Alignment, Font, Side, Border, colors
from openpyxl.formatting.rule import DataBarRule
import os
import numpy as np


def sheet_idx(row, col):
    """

    :param row:  >=1
    :param col: >=1
    :return: str
    """
    res = chr(ord('A') + col - 1) + str(row)
    return res


class Make_Report:
    # def __init__(self, save_file, model_path: str, network: str, dataset: str, train_size: int, evalute_szie: int,
    #              SNR: str, mask: str, mask_idx: str, set:str, start_row=1, sheet:int=0):
    def __init__(self, save_file, model_path: str, train_way:{}, start_row=1, sheet: int = 0):
        self.excel_path = save_file
        """ 表头测试内容 """
        self.head_row = ['mask size', 'mask idx']
        self.test_content = ['-4', '-2', '0', '2', '4', '6', '8', '10']
        self.head_row_idx = [0, 0.2, 0.4, 0.6, 0.8]

        # self.train_way1 = ['train way', 'network', network, 'train size', (train_size), 'SNR', str(SNR),
        #                    'mask idx', str(mask_idx), ' ']
        # self.train_way2 = ['train way', 'dataset', dataset, 'evaluation size', (evalute_szie), 'mask', str(mask),
        #                    'sets', set]
        self.train_way = []
        for i, k1 in enumerate(train_way.keys()):  ## 几行
            self.train_way.append(['train way'])
            for j, k2 in enumerate(train_way[k1].keys()):
                self.train_way[i].append(k2+':')
                self.train_way[i].append(train_way[k1][k2])



        self.data_buff = None
        self.data_start_row_ori = start_row
        self.data_start_col_ori = 1
        self.data_start_row = 1

        self.row_idx = {}
        self.model_path = model_path
        # save_file = r'I:\python\01-work\result\drsn_record.xlsx'
        if os.path.exists(save_file):
            self.book = oxl.load_workbook(save_file)
            print('load excel:', save_file)
            sheets = self.book.get_sheet_names()
            name = 'Sheet%d' % sheet if sheet > 0 else 'Sheet'
            if len(sheets)<(sheet):
                raise IndexError(f'your sheet idx ({sheet}) is too large. exist sheet name= [{sheets}]. idx max is {len(sheets)}')
            if sheet < 0:
                name = sheets[sheets]
            if name not in sheets :
                self.book.create_sheet(name)
                print('\tcreat sheet name=', name)
                self.data_start_row_ori = 1  ## 从头开始
                self.data_start_col_ori = 1  ## 从头开始
            else:
                print('working sheet is:', name)
            self.sheet = self.book.get_sheet_by_name(name)  # .active


        else:
            self.book = oxl.Workbook()
            print('create excel:', save_file)
            self.sheet = self.book.active
            # self.data_start_row_ori = 1  ## 从头开始
            # self.data_start_col_ori = 1  ## 从头开始

        self.__buff_init()
        self.__make_book_title(self.data_start_row_ori, self.data_start_col_ori)

    def __make_book_title(self, start_row, start_col):
        head_row = self.head_row
        test_content = self.test_content
        # train_way1 = self.train_way1
        # train_way2 = self.train_way2
        train_way = self.train_way
        model_path = self.model_path

        width = len(head_row + test_content)
        # start_row, start_col = 1, 1
        # book = self.book
        sheet = self.sheet

        alignment_center = Alignment(horizontal='center', vertical='center', )
        font_time_roman = Font(name="Times New Roman", bold=False)

        num_row = 0

        sheet.cell(start_row + num_row, start_col, 'model evaluation', )
        sheet[sheet_idx(start_row + num_row, start_col)].alignment = alignment_center
        # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        sheet.merge_cells(start_row=start_row + num_row, end_row=start_row + num_row, start_column=start_col,
                          end_column=start_col + width - 1, )

        num_row += 1
        sheet.cell(start_row + num_row, start_col, model_path, )
        sheet[sheet_idx(start_row + num_row, start_col)].alignment = Alignment(horizontal='left', vertical='center', )
        # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        sheet.merge_cells(start_row=start_row + num_row, end_row=start_row + num_row, start_column=start_col,
                          end_column=start_col + width - 1, )

        # num_row += 1
        # for i, va in enumerate(train_way1):
        #     sheet.cell(start_row + num_row, start_col + i, va, )
        #     sheet[sheet_idx(start_row + num_row, start_col + i)].alignment = alignment_center
        #     # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        # num_row += 1
        # for i, va in enumerate(train_way2):
        #     sheet.cell(start_row + num_row, start_col + i, va, )
        #     sheet[sheet_idx(start_row + num_row, start_col + i)].alignment = alignment_center
        #     # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        num = 0
        for j, train_way_idx in enumerate(train_way):
            num_row += 1
            num = j
            for i, va in enumerate(train_way_idx):
                sheet.cell(start_row + num_row, start_col + i, va, )
                sheet[sheet_idx(start_row + num_row, start_col + i)].alignment = alignment_center
                # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        print('start_row', start_row + num_row - num, 'end_row',start_row + num_row)
        sheet.merge_cells(start_row=start_row + num_row - num, end_row=start_row + num_row, start_column=start_col,
                          end_column=start_col, )  ## 'train_way'

        num_row += 1
        sheet.cell(start_row + num_row, start_col, 'SNR (dB)', )
        sheet[sheet_idx(start_row + num_row, start_col)].alignment = alignment_center
        # sheet[sheet_idx(start_row + num_row, start_col)].border = border_content
        sheet.merge_cells(start_row=start_row + num_row, end_row=start_row + num_row, start_column=start_col,
                          end_column=start_col + width - 1, )

        test_row = head_row + test_content
        num_row += 1
        for i, va in enumerate(test_row):
            sheet.cell(start_row + num_row, start_col + i, va, )
            sheet[sheet_idx(start_row + num_row, start_col + i)].alignment = alignment_center
            # sheet[sheet_idx(start_row + num_row, start_col + i)].border = border_content

        self.data_start_row = num_row + 1 + start_row

        # border_head = Border(left=Side(style='medium', color=colors.BLACK),
        #                      right=Side(style='medium', color=colors.BLACK),
        #                      top=Side(style='medium', color=colors.BLACK),
        #                      bottom=Side(style='thin', color=colors.BLACK))

        border_content = Border(left=Side(style='thin', color=colors.BLACK),
                                right=Side(style='thin', color=colors.BLACK),
                                top=Side(style='thin', color=colors.BLACK),
                                bottom=Side(style='thin', color=colors.BLACK))
        for y in range(num_row + 1):
            for x in range(width):
                sheet[sheet_idx(start_row + y, start_col + x)].font = font_time_roman
                sheet.cell(start_row + y, start_col + x).border = border_content
                # sheet[sheet_idx(start_row + y, start_col + x)].number_format =  数字格式

        # self.__filling_border('A7', 'J12')

        self.sheet = sheet

    def __filling_border(self, start_loc, end_loc, border=None):  # 参数为左上角坐标和右下角坐标，形如'D3','A5'等。ws是worksheet对象。
        x_start = start_loc[0]
        y_start = start_loc[1:len(start_loc)]  # 切片获取坐标的数字部分
        x_end = end_loc[0]
        y_end = end_loc[1:len(end_loc)]
        len_y = int(y_end) - int(y_start) + 1
        alphabet = string.ascii_uppercase  # 导入字母表
        len_x = alphabet.index(x_end) - alphabet.index(x_start) + 1
        # print( alphabet.index(x_end) , alphabet.index(x_start) , x_start,y_start, x_end, y_end, len_y, len_x  )
        ws = self.sheet

        if border is None:
            border = Side(style='thin')

        # 左上
        temp = start_loc
        # ws[temp].border = Border(left=border, top=border)
        b = copy.copy(ws[temp].border)
        b.left = border
        b.top = border
        ws[temp].border = b
        # 右下
        temp = end_loc
        # ws[temp].border = Border(right=border, bottom=border)
        b = copy.copy(ws[temp].border)
        b.right = border
        b.bottom = border
        ws[temp].border = b
        # 右上
        temp = x_end + y_start
        # ws[temp].border = Border(right=border, top=border)
        b = copy.copy(ws[temp].border)
        b.right = border
        b.top = border
        ws[temp].border = b
        # 左下
        temp = x_start + y_end
        # ws[temp].border = Border(left=border, bottom=border)
        b = copy.copy(ws[temp].border)
        b.left = border
        b.bottom = border
        ws[temp].border = b
        # 上
        for i in range(0, len_x - 2):
            temp = alphabet[alphabet.index(x_start) + 1 + i] + y_start
            # ws[temp].border = Border(top=border)
            b = copy.copy(ws[temp].border)
            b.top = border
            ws[temp].border = b
        # 下
        for i in range(0, len_x - 2):
            temp = alphabet[alphabet.index(x_start) + 1 + i] + y_end
            # ws[temp].border = Border(bottom=border)
            b = copy.copy(ws[temp].border)
            b.bottom = border
            ws[temp].border = b
        # 左
        for i in range(0, len_y - 2):
            temp = x_start + str(int(y_start) + 1 + i)
            # ws[temp].border = Border(left=border)
            b = copy.copy(ws[temp].border)
            b.left = border
            ws[temp].border = b
        # 右
        for i in range(0, len_y - 2):
            temp = x_end + str(int(y_start) + 1 + i)
            # ws[temp].border = Border(right=border)
            b = copy.copy(ws[temp].border)
            b.right = border
            ws[temp].border = b
        return 0

    def save_excel(self, save_file=None):
        if save_file is None: save_file = self.excel_path
        # save_file = r'I:\python\01-work\result\aa.xlsx'
        os.makedirs(os.path.dirname(save_file), exist_ok=True)
        self.book.save(save_file)
        # exit()

    def __buff_init(self):
        num = 0
        row_idx = {}
        for mask_size_idx in self.head_row_idx:  ## mask size
            for mask_idx_idx in self.head_row_idx:
                if mask_idx_idx + mask_size_idx > 1: break
                row_idx[str(mask_size_idx) + '_' + str(mask_idx_idx)] = num
                num += 1
                if mask_size_idx == 0: break
        row_idx['end' + '_' + 'end'] = num  ## 结束标志
        width = len(self.test_content)
        data = np.zeros([num, width], dtype=np.float)
        data[:, :] = np.nan
        # data[:3, :] = np.arange(1, 24).reshape(-1, 8).dtype(np.float)
        self.data_buff = data
        self.row_idx = row_idx

    def add_data(self, mask_size: float, mask_idx: float, snr: int, acc: float):
        col = self.test_content.index(str(snr))
        row = self.row_idx[str(mask_size) + '_' + str(mask_idx)]
        self.data_buff[row, col] = acc
        pass

    def record_data_in_excel(self):
        start_row, start_col, width = self.data_start_row, self.data_start_col_ori, len(
            self.head_row + self.test_content)
        buff = self.data_buff
        minnum = np.min(buff)
        border_content = Border(left=Side(style='thin', color=colors.BLACK),
                                right=Side(style='thin', color=colors.BLACK),
                                top=Side(style='thin', color=colors.BLACK),
                                bottom=Side(style='thin', color=colors.BLACK))

        num = 0
        last_idx = str(self.head_row_idx[0])
        for row, da in enumerate(self.row_idx.values()):  ## 有结束标志  'end'+'_'+'end'
            mask_size, mask_idx = str(list(self.row_idx.keys())[row]).split('_')
            """ 单元格合并 """
            if mask_size != last_idx:  # mask_size 自动合并
                last_idx = mask_size
                # print('==', f'start row={start_row}, row={row}, num={num},  row idx=[{start_row + row - num},'
                #             f' {start_row + row - 1}]')
                self.sheet.merge_cells(start_row=start_row + row - num, end_row=start_row + row - 1,
                                       start_column=start_col, end_column=start_col, )
                self.__filling_border(sheet_idx(start_row + row - num, start_col),
                                      sheet_idx(start_row + row - 1, start_col + width - 1))
                num = 1
            else:
                num += 1
            if mask_size == 'end': break

            """ 写入 mask_size, mask_idx 对应的数据 """
            mask_size, mask_idx = float(mask_size), float(mask_idx)
            # print('mask_size, mask_idx :',row, mask_size, mask_idx , start_row + row , mask_size)
            self.__sheet_idx_input_value(start_row + row, start_col, mask_size)
            self.__sheet_idx_input_value(start_row + row, start_col + 1, mask_idx)
            """ 写入测试内容 """
            self.data_start_row += 1
            for col, snr_idx in enumerate(self.test_content):
                dat = buff[row, col]
                if dat == np.nan: continue
                self.__sheet_idx_input_value_bar(start_row + row, 3 + col, v=dat, min=minnum, max=1)

        self.__filling_border(sheet_idx(self.data_start_row_ori, self.data_start_col_ori),
                              sheet_idx(self.data_start_row - 1, self.data_start_col_ori + width - 1),
                              border=Side(style='medium', color=colors.BLACK))

        self.save_excel()

    def __sheet_idx_input_value(self, row, col, v):
        sheet = self.sheet
        # print('__sheet_idx_input_value:',  row, col, v,)
        sheet.cell(row, col, np.round(v, 4), )
        sheet[sheet_idx(row, col)].alignment = Alignment(horizontal='center', vertical='center', )
        # sheet[sheet_idx(row, col)].number_format = '0.0000'  ## 数据显示格式 固定小数位长度
        self.sheet = sheet

    def __sheet_idx_input_value_bar(self, row, col, v, min=0, max=1):
        sheet = self.sheet
        # print('__sheet_idx_input_value:',  row, col, v,)
        sheet.cell(row, col, np.round(v, 4), )
        sheet[sheet_idx(row, col)].alignment = Alignment(horizontal='center', vertical='center', )
        # sheet[sheet_idx(row, col)].number_format = '0.00%'
        """ 添加数据进度条  """
        rule = DataBarRule(start_type="num", end_type="num", start_value=min, end_value=max,
                           color="006400"  ## RGB
                           )
        sheet.conditional_formatting.add(sheet_idx(row, col), rule)

        self.sheet = sheet
